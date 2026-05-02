"""
production.py
-------------
Generates a production plan Excel file matching the exact format
of Production__SKU.xlsx.

Filter Logic:
    A SKU is included if ANY of these are true:
    1. Amazon alert = 'out_of_stock' or 'low_stock'
    2. Amazon Recommended action = 'Create shipping plan'
    3. Days of Supply < user-defined threshold (optional)

    Excluded if BOTH of these are true:
    - units_sold_last_30_days = 0
    - recommended_replenishment_qty = 0
    (no sales and Amazon doesn't recommend restocking = skip)

Quantity Formula:
    projected_demand = units_sold_last_30_days * (coverage_weeks / 4)
    qty_needed = projected_demand - total_units_on_hand
    qty_needed = round_to_valid_qty(qty_needed)

    If qty_needed <= 0 → SKU already has enough stock → skip

Valid Quantity Set:
    Minimum = 6
    Then multiples of 12: 12, 24, 36, 48, 60, 72, 84, 96, 108, 120, 132, 144...

    round_to_valid_qty(n):
        if n <= 6  → return 6
        else       → return ceil(n / 12) * 12

Output format matches Production__SKU.xlsx exactly:
    Row 1: blank
    Row 2: No | WL | Maventee - Out of Stock -SKU | FNSKU | Quantity | Box
    Row 3+: data rows sorted by warehouse location
    Last row: total quantity sum
"""

import math
import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


DAYS_COL = "total_days_of_supply_(including_units_from_open_shipments)"


def parse_days(val) -> float:
    """Safely parse days of supply — handles '365+', NaN, etc."""
    try:
        s = str(val).replace("+", "").strip()
        return float(s)
    except (ValueError, TypeError):
        return 9999.0


def round_to_valid_qty(qty: float) -> int:
    """
    Rounds a calculated quantity to the nearest valid production quantity.

    Valid quantities: 6 (minimum), then multiples of 12.
    Examples:
        qty = 3  → 6
        qty = 8  → 12
        qty = 15 → 24
        qty = 30 → 36
        qty = 65 → 72
        qty = 144 → 144
    """
    if qty <= 6:
        return 6
    return math.ceil(qty / 12) * 12


def generate_production_plan(
    restock_df: pd.DataFrame,
    database_df: pd.DataFrame,
    coverage_weeks: int = 12,           # how many weeks of inventory to cover
    days_supply_threshold: int = None,  # optional: include if days < this
) -> bytes:
    """
    Generates a production plan Excel file.

    Args:
        restock_df:             Standardized restock report DataFrame
        database_df:            Standardized database/mapping DataFrame
        coverage_weeks:         Weeks of inventory to produce (default 12 = 3 months)
        days_supply_threshold:  Optionally include SKUs with days < this value

    Returns:
        Excel file as bytes
    """

    df = restock_df.copy()

    # ── Step 1: Normalize column names ───────────────────────────────────────
    df.columns = (df.columns.astype(str)
                  .str.strip()
                  .str.lower()
                  .str.replace(r"\s+", "_", regex=True))

    # ── Step 2: Parse numeric columns safely ─────────────────────────────────
    def safe_num(series, default=0):
        return pd.to_numeric(series, errors="coerce").fillna(default)

    df["units_sold_num"]   = safe_num(df.get("units_sold_last_30_days", pd.Series(0, index=df.index)))
    df["recommended_qty"]  = safe_num(df.get("recommended_replenishment_qty", pd.Series(0, index=df.index)))
    df["total_units_num"]  = safe_num(df.get("total_units", pd.Series(0, index=df.index)))
    df["days_supply"]      = df.get(DAYS_COL, pd.Series("9999", index=df.index)).apply(parse_days)

    # ── Step 3: Filter — include SKUs that need restocking ───────────────────
    mask_alert = df.get("alert", pd.Series("", index=df.index)).isin(
        ["out_of_stock", "low_stock"]
    )
    mask_action = (
        df.get("recommended_action", pd.Series("", index=df.index))
        .str.lower().str.strip() == "create shipping plan"
    )
    mask_days = pd.Series(False, index=df.index)
    if days_supply_threshold is not None:
        mask_days = df["days_supply"] < days_supply_threshold

    flagged = df[mask_alert | mask_action | mask_days].copy()

    # ── Step 4: Exclude SKUs with 0 sales AND 0 recommended qty ─────────────
    # Per client: "If a SKU has 0 sales last month AND no recommended quantity
    # in the report, that item should be excluded from the production run."
    # Use .astype(float) to safely handle StringDtype zeros
    units_zero = pd.to_numeric(flagged["units_sold_num"], errors="coerce").fillna(0) == 0
    rec_zero   = pd.to_numeric(flagged["recommended_qty"], errors="coerce").fillna(0) == 0
    flagged = flagged[~(units_zero & rec_zero)]

    if flagged.empty:
        raise ValueError(
            "No SKUs passed the filter. "
            "All flagged SKUs either have 0 sales and no recommended quantity, "
            "or no SKUs matched the filter criteria."
        )

    # ── Step 5: Calculate production quantity ────────────────────────────────
    # Formula: (units_sold_last_30 × coverage_weeks/4) - total_units_on_hand
    # Then round to valid quantity set (min 6, then multiples of 12)
    def calc_qty(row):
        projected_demand = row["units_sold_num"] * (coverage_weeks / 4)
        qty_needed = projected_demand - row["total_units_num"]

        if qty_needed <= 0:
            # Already have enough stock — skip this SKU
            return 0

        return round_to_valid_qty(qty_needed)

    flagged["quantity"] = flagged.apply(calc_qty, axis=1)

    # Remove SKUs where calculated quantity is 0 (already have enough stock)
    flagged = flagged[flagged["quantity"] > 0]

    if flagged.empty:
        raise ValueError(
            f"All flagged SKUs already have enough stock for {coverage_weeks} weeks. "
            f"Try reducing the coverage target or check your inventory levels."
        )

    # ── Step 6: Join with Database for Warehouse Location + Seller SKU ───────
    db = database_df.copy()
    db.columns = (db.columns.astype(str)
                  .str.strip()
                  .str.lower()
                  .str.replace(r"\s+", "_", regex=True))

    # Only pull warehouse_location from the database — we don't use seller_sku
    # final_sku always comes from merchant_sku in the restock report (Pr- prefix)
    db_slim = db[["fnsku", "warehouse_location"]].drop_duplicates(subset=["fnsku"])

    merged = flagged.merge(db_slim, on="fnsku", how="left")
    merged["warehouse_location"] = merged["warehouse_location"].fillna("—")

    # Use merchant_sku as final_sku — this preserves the Pr- prefix
    merged["final_sku"] = merged["merchant_sku"].fillna(merged["fnsku"])

    # ── Step 6b: Split into groups by SKU prefix ────────────────────────────
    # Pr-  → Amazon-fulfilled (main production plan — Order sheet)
    # P-RN → Multipacks / bundles (separate sheet)
    # Everything else (Reg-, UsaFlag1-, etc.) goes on Order sheet too
    sku_col = merged["final_sku"].astype(str)

    prn_df = merged[sku_col.str.startswith("P-RN")].copy()
    pr_df  = merged[~sku_col.str.startswith("P-RN")].copy()

    # Sort each group by warehouse location
    pr_df  = pr_df.sort_values("warehouse_location").reset_index(drop=True)
    prn_df = prn_df.sort_values("warehouse_location").reset_index(drop=True)

    # ── Step 7: Build Excel file ──────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "Order"

    # Use pr_df as the main merged for the Order sheet and Summary
    merged = pr_df

    # Build main Order sheet (Pr- SKUs only)
    _build_order_sheet(ws, merged, "Maventee - Out of Stock  -SKU")

    # ── Step 8: Summary sheet (Pr- only) ────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    _build_summary_sheet(ws2, merged, coverage_weeks)

    # ── Step 9: Multipacks sheet (P-RN- SKUs) ────────────────────────────────
    ws3 = wb.create_sheet("Multipacks")
    _build_multipack_sheet(ws3, prn_df)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _build_order_sheet(ws, df: pd.DataFrame, sku_header: str = "SKU"):
    """Builds a standard order sheet — used for Pr- and other groups."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    col_widths = [6, 12, 45, 16, 10, 8]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    header_fill  = PatternFill("solid", fgColor="4472C4")
    header_font  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    header_align = Alignment(horizontal="center", vertical="center")
    data_font    = Font(name="Arial", size=10)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align   = Alignment(horizontal="left", vertical="center")
    alt_fill     = PatternFill("solid", fgColor="F2F2F2")
    total_fill   = PatternFill("solid", fgColor="E2EFDA")
    total_font   = Font(name="Arial", bold=True, size=10)
    thin_border  = Border(bottom=Side(style="thin", color="D9D9D9"))

    # Row 1: blank
    ws.row_dimensions[1].height = 8

    # Row 2: headers
    headers = ["No", "WL", sku_header, "FNSKU", "Quantity", "Box"]
    ws.row_dimensions[2].height = 20
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = header_align

    if df.empty:
        ws.cell(row=3, column=1, value="No items in this category").font = data_font
        return

    # Data rows
    for i, (_, row) in enumerate(df.iterrows(), 1):
        r = i + 2
        ws.row_dimensions[r].height = 18

        ws.cell(row=r, column=1, value=i).alignment = center_align
        ws.cell(row=r, column=2, value=str(row["warehouse_location"])).alignment = center_align
        ws.cell(row=r, column=3, value=str(row["final_sku"])).alignment = left_align
        ws.cell(row=r, column=4, value=str(row["fnsku"])).alignment = center_align
        ws.cell(row=r, column=5, value=int(row["quantity"])).alignment = center_align
        ws.cell(row=r, column=6, value=None)

        for col in range(1, 7):
            cell = ws.cell(row=r, column=col)
            cell.font   = data_font
            cell.border = thin_border
            if i % 2 == 0:
                cell.fill = alt_fill

    # Total row
    total_row = len(df) + 3
    ws.row_dimensions[total_row].height = 20
    total_qty = int(df["quantity"].sum())

    for col in range(1, 7):
        cell = ws.cell(row=total_row, column=col)
        cell.font      = total_font
        cell.fill      = total_fill
        cell.alignment = center_align

    ws.cell(row=total_row, column=5, value=total_qty)


def _build_multipack_sheet(ws, df: pd.DataFrame):
    """
    Builds the Multipacks sheet — expands each P-RN multipack SKU into
    individual color rows. Each color in the pack gets the same quantity
    as the original (e.g. 24 units of a 6-pack = 24 of each color).
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # Color abbreviation map for multipack parsing
    MULTIPACK_COLORS = {
        "mauve": "Mauve", "mrn": "Maroon", "prp": "Purple",
        "slt": "Slate", "lgn": "Lagoon Blue", "lg": "Light Gray",
        "blk": "Black", "mgrn": "Military Green", "nvy": "Navy",
        "dg": "Dark Gray", "forest": "Forest Green",
        "red": "Red", "royal": "Royal Blue",
        "black": "Black", "navy": "Navy", "slate": "Slate",
        "lagoon": "Lagoon Blue", "purple": "Purple", "maroon": "Maroon",
        "dgray": "Dark Gray", "lgray": "Light Gray",
    }

    SIZE_NORMALIZE = {
        "2xlarge": "2XL", "3xlarge": "3XL", "xlarge": "XL",
        "x-large": "XL", "large": "L", "medium": "M", "small": "S",
        "xxl": "2XL",
    }

    def parse_multipack_colors(sku):
        """Extract individual colors from a multipack SKU."""
        sku = str(sku)
        # Get the color segment — between last two dashes
        parts = sku.split("-")
        if len(parts) < 2:
            return ["Unknown"]
        color_seg = parts[-2]  # e.g. "Mauve/Mrn/Prp/Slt/Lgn/Lg" or "(Navy-Slate-MGreen)"
        # Remove parentheses
        color_seg = color_seg.replace("(", "").replace(")", "")
        # Split by / or internal -
        if "/" in color_seg:
            raw_colors = color_seg.split("/")
        else:
            raw_colors = color_seg.split("-")
        # Normalize each color
        result = []
        for c in raw_colors:
            key = c.strip().lower().replace(" ", "")
            result.append(MULTIPACK_COLORS.get(key, c.strip()))
        return result

    col_widths = [6, 12, 25, 20, 16, 10, 8]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    header_fill  = PatternFill("solid", fgColor="1e2468")
    header_font  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    header_align = Alignment(horizontal="center", vertical="center")
    data_font    = Font(name="Arial", size=10)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align   = Alignment(horizontal="left", vertical="center")
    alt_fill     = PatternFill("solid", fgColor="F2F2F2")
    total_fill   = PatternFill("solid", fgColor="E2EFDA")
    total_font   = Font(name="Arial", bold=True, size=10)
    thin_border  = Border(bottom=Side(style="thin", color="D9D9D9"))

    # Row 1: blank
    ws.row_dimensions[1].height = 8

    # Row 2: headers
    headers = ["No", "WL", "Original SKU", "Color", "Size", "Quantity", "Box"]
    ws.row_dimensions[2].height = 20
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = header_align

    if df.empty:
        ws.cell(row=3, column=1, value="No multipack items in this production run").font = data_font
        return

    # Expand each multipack into individual color rows
    expanded_rows = []
    for _, row in df.iterrows():
        sku = str(row["final_sku"])
        qty = int(row["quantity"])
        wl  = str(row["warehouse_location"])

        # Parse size
        parts = sku.split("-")
        raw_size = parts[-1].strip()
        size = SIZE_NORMALIZE.get(raw_size.lower().replace(" ", ""), raw_size)

        # Parse individual colors
        colors = parse_multipack_colors(sku)

        for color in colors:
            expanded_rows.append({
                "warehouse_location": wl,
                "original_sku": sku,
                "color": color,
                "size": size,
                "quantity": qty,
            })

    # Write expanded rows
    row_num = 3
    total_qty = 0
    for i, erow in enumerate(expanded_rows, 1):
        ws.row_dimensions[row_num].height = 18

        ws.cell(row=row_num, column=1, value=i).alignment = center_align
        ws.cell(row=row_num, column=2, value=erow["warehouse_location"]).alignment = center_align
        ws.cell(row=row_num, column=3, value=erow["original_sku"]).alignment = left_align
        ws.cell(row=row_num, column=4, value=erow["color"]).alignment = left_align
        ws.cell(row=row_num, column=5, value=erow["size"]).alignment = center_align
        ws.cell(row=row_num, column=6, value=erow["quantity"]).alignment = center_align
        ws.cell(row=row_num, column=7, value=None)

        for col in range(1, 8):
            cell = ws.cell(row=row_num, column=col)
            cell.font   = data_font
            cell.border = thin_border
            if i % 2 == 0:
                cell.fill = alt_fill

        total_qty += erow["quantity"]
        row_num += 1

    # Total row
    ws.row_dimensions[row_num].height = 20
    for col in range(1, 8):
        cell = ws.cell(row=row_num, column=col)
        cell.font      = total_font
        cell.fill      = total_fill
        cell.alignment = center_align
    ws.cell(row=row_num, column=1, value="TOTAL")
    ws.cell(row=row_num, column=6, value=total_qty)


def _build_summary_sheet(ws, merged: pd.DataFrame, coverage_weeks: int):
    """
    Summary sheet showing:
    - Production run metadata (date, coverage weeks, total SKUs, total units)
    - Blank shirts needed by color
    - Blank shirts needed by size
    """
    from datetime import date

    if merged.empty:
        ws.cell(row=1, column=1, value="No Pr- SKUs in this production run")
        return

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 16

    navy_fill   = PatternFill("solid", fgColor="1e2468")
    orange_fill = PatternFill("solid", fgColor="E07B2A")
    alt_fill    = PatternFill("solid", fgColor="F2F2F2")
    total_fill  = PatternFill("solid", fgColor="E2EFDA")

    white_bold  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    bold_font   = Font(name="Arial", bold=True, size=10)
    data_font   = Font(name="Arial", size=10)
    center      = Alignment(horizontal="center", vertical="center")
    left        = Alignment(horizontal="left", vertical="center")

    # ── Run metadata ─────────────────────────────────────────────────────────
    row = 1
    ws.cell(row=row, column=1, value="Production Run Summary").font = Font(name="Arial", bold=True, color="FFFFFF", size=12)
    ws.cell(row=row, column=1).fill = navy_fill
    ws.cell(row=row, column=1).alignment = left
    ws.merge_cells(f"A{row}:C{row}")
    ws.row_dimensions[row].height = 22
    row += 1

    meta = [
        ("Generated", str(date.today())),
        ("Coverage Target", f"{coverage_weeks} weeks"),
        ("Total SKUs", str(len(merged))),
        ("Total Units to Produce", str(int(merged["quantity"].sum()))),
    ]
    for label, value in meta:
        ws.cell(row=row, column=1, value=label).font = bold_font
        ws.cell(row=row, column=2, value=value).font = data_font
        row += 1

    row += 1  # spacer

    # ── Parse color and size from SKU ────────────────────────────────────────

    # Known garment types — second to last segment when present
    GARMENT_TYPES = {
        "toddler", "youth", "unisex", "uni", "onesie",
        "babyt", "(babyt)", "ma", "dtg", "dtg_ma",
    }

    # Size normalization — full words to abbreviations
    SIZE_NORMALIZE = {
        "2xlarge": "2XL", "3xlarge": "3XL", "4xlarge": "4XL",
        "xlarge": "XL", "x-large": "XL",
        "large": "L", "medium": "M", "small": "S",
        "2xsmall": "2XS", "xsmall": "XS", "x-small": "XS",
        "xxl": "2XL", "xxxl": "3XL",
        "12m": "12m", "18m": "18m", "24m": "24m",
    }

    # Color normalization — abbreviations and variants to full names
    COLOR_NORMALIZE = {
        "bl": "Black",
        "black": "Black",
        "ma": "Mauve",
        "b.ma": "Mauve",
        "bma": "Mauve",
        "dtg_ma": "Mauve",
        "dtgma": "Mauve",
        "mauve": "Mauve",
        "dg": "Dark Gray",
        "dgray": "Dark Gray",
        "darkgray": "Dark Gray",
        "lg": "Light Gray",
        "lgray": "Light Gray",
        "lightgray": "Light Gray",
        "fgreen": "Forest Green",
        "f.green": "Forest Green",
        "forestgreen": "Forest Green",
        "mgreen": "Military Green",
        "militarygreen": "Military Green",
        "hnavy": "Heather Navy",
        "heathernavy": "Heather Navy",
        "hpurple": "Heather Purple",
        "navy": "Navy",
        "white": "White",
        "w": "White",
        "natural": "Natural",
        "red": "Red",
        "royal": "Royal Blue",
        "royalblue": "Royal Blue",
        "iblue": "Indigo Blue",
        "indigoblue": "Indigo Blue",
        "lagoon": "Lagoon Blue",
        "lagoonblue": "Lagoon Blue",
        "brown": "Brown",
        "maroon": "Maroon",
        "purple": "Purple",
        "slate": "Slate",
        "yellow": "Yellow",
        "orange": "Orange",
        "texorange": "Texas Orange",
        "lavender": "Lavender",
        "mgreen": "Military Green",
    }

    def extract_color_size(sku):
        sku = str(sku)

        # Edge case 1: Multipack SKUs with slash colors or P-RN bundle format
        # Label as "Multipack (assorted)" so they dont pollute the summary
        if "/" in sku or ("P-RN" in sku and any(f"-{n}P-" in sku for n in ["1","2","3","4","6","12"])):
            parts = sku.split("-")
            raw_size = parts[-1].strip()
            size_key = raw_size.lower().replace(" ", "")
            size = SIZE_NORMALIZE.get(size_key, raw_size)
            return "Multipack (assorted)", size

        # Edge case 2: SKU too short to parse
        parts = [p.strip() for p in sku.split("-")]
        if len(parts) < 2:
            return "Unknown", "Unknown"

        raw_size     = parts[-1]
        raw_garment  = parts[-2].lower().replace("(", "").replace(")", "")
        raw_color    = parts[-3] if len(parts) >= 3 else "Unknown"

        # Normalize size
        size_key = raw_size.lower().replace(" ", "")
        size = SIZE_NORMALIZE.get(size_key, raw_size)

        # Determine garment type and full size label
        if raw_garment in GARMENT_TYPES or raw_garment.startswith("onesie"):
            if raw_garment in ("uni", "unisex"):
                # Unisex = default adult size — no prefix needed
                # "Unisex L" is just "L", distinction only needed for Youth/Toddler etc.
                full_size = size
            elif raw_garment in ("babyt", "(babyt)"):
                full_size = f"Baby T {size}"
            elif raw_garment in ("dtg", "dtg_ma", "ma"):
                # Print codes not garment types — no prefix, fix color
                raw_color = parts[-3] if len(parts) >= 3 else raw_color
                full_size = size
            elif raw_garment.startswith("onesie"):
                full_size = f"Onesie {size}"
            else:
                # Toddler, Youth, etc. — keep the prefix for distinction
                full_size = f"{raw_garment.capitalize()} {size}"
        else:
            # No garment type found — second to last is likely color
            raw_color = raw_garment
            full_size = size

        # Normalize color
        color_key = raw_color.lower().replace(" ", "").replace("_", "")
        # handle b.ma → mauve before stripping dots
        color_key_nodot = color_key.replace(".", "")
        color = COLOR_NORMALIZE.get(color_key, COLOR_NORMALIZE.get(color_key_nodot, raw_color))

        return color, full_size

    if merged.empty:
        merged["color"] = pd.Series(dtype="object")
        merged["size"]  = pd.Series(dtype="object")
    else:
        result = merged["final_sku"].apply(
            lambda s: pd.Series(extract_color_size(s), index=["color", "size"])
        )
        merged[["color", "size"]] = result

    # ── Color + Size combination summary ─────────────────────────────────────
    # Shows total units needed per color+size combo so warehouse knows
    # exactly how many of each blank to pull (e.g. 48 Black Unisex 2XL)
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 16

    ws.cell(row=row, column=1, value="Blank Shirts Needed — Color & Size").font = white_bold
    ws.cell(row=row, column=1).fill = orange_fill
    ws.merge_cells(f"A{row}:C{row}")
    ws.row_dimensions[row].height = 20
    row += 1

    for label, col in [("Color", 1), ("Size", 2), ("Total Units", 3)]:
        ws.cell(row=row, column=col, value=label).font = white_bold
        ws.cell(row=row, column=col).fill = navy_fill
        ws.cell(row=row, column=col).alignment = center
    row += 1

    # Group by color + size, sort by color then size
    combo_summary = (
        merged.groupby(["color", "size"])["quantity"]
        .sum()
        .reset_index()
        .sort_values(["color", "size"])
    )

    current_color = None
    for i, combo_row in enumerate(combo_summary.itertuples()):
        # Shade alternating color groups for readability
        fill = alt_fill if i % 2 == 1 else None

        ws.cell(row=row, column=1, value=combo_row.color).font = data_font
        ws.cell(row=row, column=2, value=combo_row.size).font = data_font
        ws.cell(row=row, column=3, value=int(combo_row.quantity)).font = data_font
        ws.cell(row=row, column=3).alignment = center

        if fill:
            for col in [1, 2, 3]:
                ws.cell(row=row, column=col).fill = fill

        # Bold the color name only when it changes (group header effect)
        if combo_row.color != current_color:
            ws.cell(row=row, column=1).font = bold_font
            current_color = combo_row.color

        row += 1

    # Total row
    ws.cell(row=row, column=1, value="TOTAL").font = bold_font
    ws.cell(row=row, column=3, value=int(merged["quantity"].sum())).font = bold_font
    ws.cell(row=row, column=3).alignment = center
    for col in [1, 2, 3]:
        ws.cell(row=row, column=col).fill = total_fill
    row += 2

    # ── Subtotals by color (quick reference) ─────────────────────────────────
    ws.cell(row=row, column=1, value="Subtotal by Color").font = white_bold
    ws.cell(row=row, column=1).fill = orange_fill
    ws.merge_cells(f"A{row}:C{row}")
    ws.row_dimensions[row].height = 20
    row += 1

    for label, col in [("Color", 1), ("Total Units", 2)]:
        ws.cell(row=row, column=col, value=label).font = white_bold
        ws.cell(row=row, column=col).fill = navy_fill
        ws.cell(row=row, column=col).alignment = center
    row += 1

    color_totals = merged.groupby("color")["quantity"].sum().sort_values(ascending=False)
    for i, (color, qty) in enumerate(color_totals.items()):
        ws.cell(row=row, column=1, value=color).font = data_font
        ws.cell(row=row, column=2, value=int(qty)).font = data_font
        ws.cell(row=row, column=2).alignment = center
        if i % 2 == 1:
            for col in [1, 2]:
                ws.cell(row=row, column=col).fill = alt_fill
        row += 1

    ws.cell(row=row, column=1, value="TOTAL").font = bold_font
    ws.cell(row=row, column=2, value=int(merged["quantity"].sum())).font = bold_font
    ws.cell(row=row, column=2).alignment = center
    for col in [1, 2]:
        ws.cell(row=row, column=col).fill = total_fill


def calculate_production_items(
    restock_df: pd.DataFrame,
    database_df: pd.DataFrame,
    coverage_weeks: int = 12,
    days_supply_threshold: int = None,
) -> list:
    """
    Runs the full production plan filter + quantity logic but returns
    a list of dicts instead of an Excel file.
    Used by the preview/batching system.

    Returns list of dicts with:
        merchant_sku, fnsku, warehouse_location, quantity, color, size
    """
    df = restock_df.copy()
    df.columns = (df.columns.astype(str)
                  .str.strip()
                  .str.lower()
                  .str.replace(r"\s+", "_", regex=True))

    def safe_num(series, default=0):
        return pd.to_numeric(series, errors="coerce").fillna(default)

    df["units_sold_num"]  = safe_num(df.get("units_sold_last_30_days", pd.Series(0, index=df.index)))
    df["recommended_qty"] = safe_num(df.get("recommended_replenishment_qty", pd.Series(0, index=df.index)))
    df["total_units_num"] = safe_num(df.get("total_units", pd.Series(0, index=df.index)))
    df["days_supply"]     = df.get(DAYS_COL, pd.Series("9999", index=df.index)).apply(parse_days)

    mask_alert  = df.get("alert", pd.Series("", index=df.index)).isin(["out_of_stock", "low_stock"])
    mask_action = (df.get("recommended_action", pd.Series("", index=df.index))
                   .str.lower().str.strip() == "create shipping plan")
    mask_days   = pd.Series(False, index=df.index)
    if days_supply_threshold is not None:
        mask_days = df["days_supply"] < days_supply_threshold

    flagged = df[mask_alert | mask_action | mask_days].copy()

    units_zero = pd.to_numeric(flagged["units_sold_num"], errors="coerce").fillna(0) == 0
    rec_zero   = pd.to_numeric(flagged["recommended_qty"], errors="coerce").fillna(0) == 0
    flagged = flagged[~(units_zero & rec_zero)]

    if flagged.empty:
        raise ValueError("No SKUs passed the filter.")

    def calc_qty(row):
        projected = row["units_sold_num"] * (coverage_weeks / 4)
        needed = projected - row["total_units_num"]
        if needed <= 0:
            return 0
        return round_to_valid_qty(needed)

    flagged["quantity"] = flagged.apply(calc_qty, axis=1)
    flagged = flagged[flagged["quantity"] > 0]

    if flagged.empty:
        raise ValueError(f"All flagged SKUs already have enough stock for {coverage_weeks} weeks.")

    # Join with database for warehouse location
    db = database_df.copy()
    db.columns = (db.columns.astype(str)
                  .str.strip()
                  .str.lower()
                  .str.replace(r"\s+", "_", regex=True))

    db_slim = db[["fnsku", "warehouse_location"]].drop_duplicates(subset=["fnsku"])
    merged  = flagged.merge(db_slim, on="fnsku", how="left")
    merged["warehouse_location"] = merged["warehouse_location"].fillna("—")
    merged["final_sku"] = merged["merchant_sku"].fillna(merged["fnsku"])

    # Exclude P-RN multipacks from main items (they get separate handling)
    sku_col = merged["final_sku"].astype(str)
    merged  = merged[~sku_col.str.startswith("P-RN")].copy()
    merged  = merged.sort_values("warehouse_location").reset_index(drop=True)

    # Parse color and size
    GARMENT_TYPES = {
        "toddler", "youth", "unisex", "uni", "onesie",
        "babyt", "(babyt)", "ma", "dtg", "dtg_ma",
    }
    SIZE_NORMALIZE = {
        "2xlarge": "2XL", "3xlarge": "3XL", "4xlarge": "4XL",
        "xlarge": "XL", "x-large": "XL",
        "large": "L", "medium": "M", "small": "S",
        "2xsmall": "2XS", "xsmall": "XS", "x-small": "XS",
        "xxl": "2XL", "xxxl": "3XL",
    }
    COLOR_NORMALIZE = {
        "bl": "Black", "black": "Black",
        "ma": "Mauve", "b.ma": "Mauve", "bma": "Mauve",
        "dtg_ma": "Mauve", "dtgma": "Mauve", "mauve": "Mauve",
        "dg": "Dark Gray", "dgray": "Dark Gray", "darkgray": "Dark Gray",
        "lg": "Light Gray", "lgray": "Light Gray", "lightgray": "Light Gray",
        "fgreen": "Forest Green", "forestgreen": "Forest Green",
        "mgreen": "Military Green", "militarygreen": "Military Green",
        "hnavy": "Heather Navy", "heathernavy": "Heather Navy",
        "navy": "Navy", "white": "White", "w": "White",
        "natural": "Natural", "red": "Red",
        "royal": "Royal Blue", "royalblue": "Royal Blue",
        "iblue": "Indigo Blue", "lagoon": "Lagoon Blue", "lagoonblue": "Lagoon Blue",
        "brown": "Brown", "maroon": "Maroon", "purple": "Purple",
        "slate": "Slate", "yellow": "Yellow", "texorange": "Texas Orange",
    }

    def extract_cs(sku):
        sku = str(sku)
        if "/" in sku:
            parts = sku.split("-")
            raw_size = parts[-1].strip()
            size = SIZE_NORMALIZE.get(raw_size.lower().replace(" ", ""), raw_size)
            return "Multipack (assorted)", size
        parts = [p.strip() for p in sku.split("-")]
        if len(parts) < 2:
            return "Unknown", "Unknown"
        raw_size    = parts[-1]
        raw_garment = parts[-2].lower().replace("(", "").replace(")", "")
        raw_color   = parts[-3] if len(parts) >= 3 else "Unknown"
        size_key = raw_size.lower().replace(" ", "")
        size = SIZE_NORMALIZE.get(size_key, raw_size)
        if raw_garment in GARMENT_TYPES or raw_garment.startswith("onesie"):
            if raw_garment in ("uni", "unisex"):
                full_size = size
            elif raw_garment in ("babyt", "(babyt)"):
                full_size = f"Baby T {size}"
            elif raw_garment in ("dtg", "dtg_ma", "ma"):
                raw_color = parts[-3] if len(parts) >= 3 else raw_color
                full_size = size
            elif raw_garment.startswith("onesie"):
                full_size = f"Onesie {size}"
            else:
                full_size = f"{raw_garment.capitalize()} {size}"
        else:
            raw_color = raw_garment
            full_size = size
        color_key = raw_color.lower().replace(" ", "").replace("_", "")
        color_key_nodot = color_key.replace(".", "")
        color = COLOR_NORMALIZE.get(color_key, COLOR_NORMALIZE.get(color_key_nodot, raw_color))
        return color, full_size

    items = []
    for _, row in merged.iterrows():
        color, size = extract_cs(str(row["final_sku"]))
        items.append({
            "merchant_sku":       str(row["final_sku"]),
            "fnsku":              str(row.get("fnsku", "")),
            "warehouse_location": str(row["warehouse_location"]),
            "quantity":           int(row["quantity"]),
            "color":              color,
            "size":               size,
        })

    return items


def generate_batch_excel(items: list, run_name: str, batch_name: str) -> bytes:
    """
    Generates a receipt Excel file for a single batch.

    Uses actual_qty if available (set by warehouse), otherwise falls back to planned quantity.
    If box_splits exist for an item, expands into one row per box split.
    items = list of dicts from batch_items table, may include box_splits.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Order"

    # ── Build expanded rows (one per box split, or one per SKU if no splits) ──
    expanded_rows = []
    for item in items:
        # Use actual_qty if set, otherwise fall back to planned quantity
        final_qty = item.get("actual_qty") if item.get("actual_qty") is not None else item["quantity"]
        splits    = item.get("box_splits", [])

        if splits:
            # One row per box split
            for split in splits:
                expanded_rows.append({
                    "final_sku":           item["merchant_sku"],
                    "warehouse_location":  item.get("warehouse_location", "—"),
                    "fnsku":               item.get("fnsku", ""),
                    "quantity":            int(split["box_qty"]),
                    "box_number":          split["box_number"],
                    "color":               item.get("color", ""),
                    "size":                item.get("size", ""),
                })
        else:
            # No box splits — single row with full actual qty
            expanded_rows.append({
                "final_sku":           item["merchant_sku"],
                "warehouse_location":  item.get("warehouse_location", "—"),
                "fnsku":               item.get("fnsku", ""),
                "quantity":            int(final_qty),
                "box_number":          item.get("box_number"),
                "color":               item.get("color", ""),
                "size":                item.get("size", ""),
            })

    df_expanded = pd.DataFrame(expanded_rows)

    # ── Order sheet ──────────────────────────────────────────────────────────
    _build_order_sheet(ws, df_expanded, f"{run_name} - {batch_name} - RECEIPT")

    # ── Summary sheet (uses actual totals per color/size) ────────────────────
    ws2 = wb.create_sheet("Summary")
    # Collapse box splits back to per-SKU for summary
    summary_rows = []
    for item in items:
        final_qty = item.get("actual_qty") if item.get("actual_qty") is not None else item["quantity"]
        summary_rows.append({
            "final_sku": item["merchant_sku"],
            "quantity":  int(final_qty),
            "color":     item.get("color", ""),
            "size":      item.get("size", ""),
        })
    df_summary = pd.DataFrame(summary_rows)
    _build_summary_sheet(ws2, df_summary, 0)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()